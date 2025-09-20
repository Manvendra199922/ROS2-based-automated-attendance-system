import rclpy
from rclpy.node import Node
from sensor_msgs.msg import Image
from cv_bridge import CvBridge
import cv2
import face_recognition
from openpyxl import Workbook, load_workbook
import os
import numpy as np

class AttendanceSystemNode(Node):
    def __init__(self):
        super().__init__('attendance_system_node')
        
        # Camera setup
        self.camera_publisher_ = self.create_publisher(Image, 'camera_image', 10)
        self.bridge = CvBridge()
        self.cap = cv2.VideoCapture(0)
        self.camera_timer = self.create_timer(0.1, self.camera_timer_callback)  # 10 Hz
        
        # Face detection setup
        haarcascade_path = '/home/manvendra/ros2_ws/haarcascade_frontalface_default.xml'
        self.face_cascade = cv2.CascadeClassifier(haarcascade_path)
        self.image_path = '/home/manvendra/ros2_ws/face_detected/face_detected_image.png'
        
        # Batch image publisher setup
        self.batch_publisher_ = self.create_publisher(Image, 'image_topic', 10)
        self.test_image_path = self.image_path
        self.folder_paths = [
            '/home/manvendra/ros2_ws/student/manvendra',
            '/home/manvendra/ros2_ws/student/shubham',
            '/home/manvendra/ros2_ws/student/bhavya',
            # Add more paths to student folders here
        ]
        
        # Excel setup
        self.excel_path = '/home/manvendra/ros2_ws/attendance.xlsx'
        self.setup_workbook()

        self.batch_images_ready = False  # Flag to check if images have been processed and are ready to publish
        self.publish_count = 0  # Counter for publishing images

    def camera_timer_callback(self):
        if not self.batch_images_ready:
            ret, frame = self.cap.read()
            if ret:
                msg = self.bridge.cv2_to_imgmsg(frame, 'bgr8')
                self.camera_publisher_.publish(msg)

                # Perform face detection
                gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                faces = self.face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

                if len(faces) > 0:
                    if os.path.exists(self.image_path):
                        os.remove(self.image_path)

                    cv2.imwrite(self.image_path, frame)
                    self.get_logger().info(f'Face detected and image saved at {self.image_path}')

                    # Perform anti-spoofing check
                    is_live = self.anti_spoofing_check(frame, faces)
                    if is_live:
                        self.get_logger().info("Live face detected")
                        # Indicate that the face detection is complete and the batch images are ready
                        self.batch_images_ready = True
                        self.publish_images_batch()
                    else:
                        self.get_logger().warn("Abnormal behaviour detected!")
                        rclpy.shutdown()

    def publish_images_batch(self):
        while self.publish_count < 3:
            test_image = cv2.imread(self.test_image_path)
            images_data = {'test_image': test_image, 'students': {}}

            # Add all student images to the batch
            for folder_path in self.folder_paths:
                student_name = os.path.basename(folder_path)
                images_data['students'][student_name] = []
                for image_name in os.listdir(folder_path):
                    image_path = os.path.join(folder_path, image_name)
                    image = cv2.imread(image_path)
                    if image is not None:
                        images_data['students'][student_name].append(image)

            # Convert images to ROS Image messages and publish
            for student_name, images in images_data['students'].items():
                for image in images:
                    msg = self.bridge.cv2_to_imgmsg(image, "bgr8")
                    msg.header.frame_id = student_name
                    self.batch_publisher_.publish(msg)
            
            # Publish the test image last
            test_msg = self.bridge.cv2_to_imgmsg(images_data['test_image'], "bgr8")
            test_msg.header.frame_id = "test_image"
            self.batch_publisher_.publish(test_msg)

            self.publish_count += 1

    def anti_spoofing_check(self, image, faces):
        """
        Basic anti-spoofing check based on color distribution.
        """
        for (x, y, w, h) in faces:
            face_roi = image[y:y+h, x:x+w]
            hsv = cv2.cvtColor(face_roi, cv2.COLOR_BGR2HSV)
            h, s, v = cv2.split(hsv)
            
            # Check the color variance
            if np.var(h) < 10 or np.var(s) < 10 or np.var(v) < 10:
                return False
        return True

    def setup_workbook(self):
        try:
            if os.path.exists(self.excel_path):
                # Load existing workbook
                self.workbook = load_workbook(self.excel_path)
                self.sheet = self.workbook.active
            else:
                # Create a new workbook and add headers
                self.workbook = Workbook()
                self.sheet = self.workbook.active
                self.sheet.append(['Timestamp', 'Date', 'Student Name', 'Match Score'])
                self.workbook.save(self.excel_path)
        except Exception as e:
            self.get_logger().error(f"Error setting up workbook: {e}")
            rclpy.shutdown()

def main(args=None):
    rclpy.init(args=args)
    attendance_system_node = AttendanceSystemNode()

    try:
        rclpy.spin(attendance_system_node)
    except KeyboardInterrupt:
        pass

    attendance_system_node.destroy_node()
    rclpy.shutdown()

if __name__ == '__main__':
    main()

