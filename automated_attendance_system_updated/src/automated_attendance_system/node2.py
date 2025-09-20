import rclpy  
from rclpy.node import Node
from sensor_msgs.msg import Image
from cv_bridge import CvBridge
import cv2
import face_recognition
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

class BatchImageSubscriber(Node):

    def __init__(self):
        super().__init__('batch_image_subscriber')
        self.subscription = self.create_subscription(
            Image,
            'image_topic',
            self.listener_callback,
            10)
        self.bridge = CvBridge()
        self.test_image_encoding = None
        self.results = {}

        # Path to the Excel file
        self.excel_path = '/home/manvendra/ros2_ws/attendance.xlsx'
        
        # Try to load an existing workbook, or create a new one if it doesn't exist
        self.setup_workbook()

    def setup_workbook(self):
        try:
            if os.path.exists(self.excel_path):
                # Load existing workbook
                self.workbook = load_workbook(self.excel_path)
                self.sheet = self.workbook.active
            else:
                # Create a new workbook and add headers
                self.workbook = Workbook()
                self.sheet = self.workbook.active
                self.sheet.append(['Timestamp', 'Date', 'Student Name', 'Match Score'])
                self.workbook.save(self.excel_path)
        except Exception as e:
            self.get_logger().error(f"Error setting up workbook: {e}")
            rclpy.shutdown()

    def listener_callback(self, msg):
        cv_image = self.bridge.imgmsg_to_cv2(msg, 'bgr8')

        if msg.header.frame_id == "test_image":
            # The test image is already processed for face detection and saved by the dedicated face detection node.
            self.test_image_encoding = self.get_face_encoding(cv_image)
            self.evaluate_images()
            rclpy.shutdown()
        else:
            student_name = msg.header.frame_id
            student_encoding = self.get_face_encoding(cv_image)
            self.results[student_name] = student_encoding

    def get_face_encoding(self, image):
        """
        Extracts face encoding from the given image. If no face is detected, returns None.
        """
        # Convert the image from BGR to RGB
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        
        # Detect faces and extract face encodings
        face_locations = face_recognition.face_locations(rgb_image)
        encodings = face_recognition.face_encodings(rgb_image, face_locations)

        if encodings:
            return encodings[0]  # Return the first face encoding found
        else:
            return None  # No face detected

    def evaluate_images(self):
        if self.test_image_encoding is None:
            self.get_logger().warn("Test image not received or no face detected!")
            return

        best_score = 0
        best_student = "Not found"
        for student_name, student_encoding in self.results.items():
            if student_encoding is not None:
                # Compare the face encodings and get the face distance
                face_distances = face_recognition.face_distance([self.test_image_encoding], student_encoding)
                match_score = 1 - face_distances[0]
                
                if match_score > best_score:
                    best_score = match_score
                    best_student = student_name

        if best_score > 0.60:  # Adjust threshold as needed
            self.update_spreadsheet(best_student, best_score)
            self.get_logger().info(f"Best match: {best_student} with Match Score: {best_score:.4f}")
        else:
            self.get_logger().info("Student not found, try again.")

    def update_spreadsheet(self, student_name, match_score):
        current_time = datetime.now().strftime('%H:%M:%S')
        current_date = datetime.now().strftime('%Y-%m-%d')

        self.sheet.append([current_time, current_date, student_name, f"{match_score:.4f}"])
        try:
            self.workbook.save(self.excel_path)
        except Exception as e:
            self.get_logger().error(f"Error saving workbook: {e}")

def main(args=None):
    rclpy.init(args=args)
    batch_image_subscriber = BatchImageSubscriber()
    rclpy.spin(batch_image_subscriber)
    batch_image_subscriber.destroy_node()
    rclpy.shutdown()

if __name__ == '__main__':
    main()
